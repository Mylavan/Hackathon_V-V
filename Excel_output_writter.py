from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import openpyxl
import os


def safe_delete_excel(file_path):
    try:
        # Try opening with read+write permissions
        with open(file_path, "r+b"):
            pass
    except PermissionError:
        print(f"❌ File '{file_path}' is open in Excel. Please close it first.")
        return False  # can't delete

    # If we got here, file is not locked
    if os.path.exists(file_path):
        os.remove(file_path)
        print(f"✅ File '{file_path}' deleted.")
        return True
    return False
# Utility to clean cell value for text search
def clean_cell_value(cell):
    # Converts lists to comma-separated string, returns empty if None
    if isinstance(cell, list):
        return ", ".join(str(x) for x in cell)
    return str(cell) if cell is not None else ""

# Utility to clean and uppercase cell value for transducer/system search
def clean_cell_value_upper(cell):
    # Uppercase string, strips brackets and spaces
    if cell is None:
        return ""
    return str(cell).replace("[", "").replace("]", "").strip().upper()

# Utility to get or create required worksheet
def get_or_create_sheet(wb, sheet_number, title_prefix="Sheet"):
    # Ensures sheet_number-th sheet exists
    while len(wb.worksheets) < sheet_number:
        wb.create_sheet(title=f"{title_prefix}{len(wb.worksheets)+1}")
    return wb.worksheets[sheet_number - 1]

# Utility to write headers if not present
def ensure_headers(ws, headers):
    # Writes headers in row 1 if empty
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        for col, key in enumerate(headers, start=1):
            ws.cell(row=1, column=col, value=key)

import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime

def normalize_date(val):
    """Try to normalize value into a date object for comparison."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()  # strip time
    if isinstance(val, str):
        for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y"):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return val  # return original if not a date

def compare_rows_sheet1_sheet2(file_path, sheet_number, row1, row2, start_col, last_col):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[sheet_number - 1]

    green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for col in range(start_col, last_col + 1):
        val1 = ws.cell(row=row1, column=col).value
        val2 = ws.cell(row=row2, column=col).value

        # Normalize for comparison
        norm1 = normalize_date(val1)
        norm2 = normalize_date(val2)

        ws.cell(row=row2, column=col).fill = green_fill if norm1 == norm2 else red_fill

    wb.save(file_path)

# Saves two result dicts as two rows in Excel, writes headers if needed
def save_results_to_excel(Excel_file_name, final_keys, TestData_Results, TestProcedure_Results, start_row=2, sheet_number=1):
    if os.path.exists(Excel_file_name):
        wb = load_workbook(Excel_file_name)
    else:
        wb = Workbook()
    ws = get_or_create_sheet(wb, sheet_number)
    ensure_headers(ws, final_keys)
    for col, key in enumerate(final_keys, start=1):
        ws.cell(row=start_row, column=col, value=str(TestData_Results.get(key, "")))
        ws.cell(row=start_row + 1, column=col, value=str(TestProcedure_Results.get(key, "")))
    wb.save(Excel_file_name)

# Saves a single dict result as a row in Excel, writes headers if needed
def save_list_to_excel_single(Excel_file_name, final_keys, Results, start_row=2, sheet_number=1):
    if os.path.exists(Excel_file_name):
        wb = load_workbook(Excel_file_name)
    else:
        wb = Workbook()
    ws = get_or_create_sheet(wb, sheet_number)
    ensure_headers(ws, final_keys)
    for col, key in enumerate(final_keys, start=1):
        ws.cell(row=start_row, column=col, value=str(Results.get(key, "")))
    wb.save(Excel_file_name)

# Writes list as a string to a specific cell in Excel
def write_list_to_cell(excel_file, sheet_number, row, col, data_list):
    content = str(data_list)
    if os.path.exists(excel_file):
        wb = load_workbook(excel_file)
    else:
        wb = Workbook()
    ws = get_or_create_sheet(wb, sheet_number)
    ws.cell(row=row, column=col, value=content)
    wb.save(excel_file)

# Searches for three strings in a row anywhere in Excel (case-sensitive)
def search_Plan_excel(file_path, str1, str2, str3):
    def clean_str(s):
        return s.replace('[', '').replace(']', '').replace("'", "").replace('"', '').strip() if s else ""
    str1, str2, str3 = map(clean_str, (str1, str2, str3))
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        row_values = [clean_cell_value(cell) for cell in row if cell is not None]
        if (any(str1 in val for val in row_values) and
            any(str2 in val for val in row_values) and
            any(str3 in val for val in row_values)):
            wb.close()
            return True
    wb.close()
    return False

# Searches for up to three strings in all rows, returns all matching rows
def search_excel_for_three_strings(file_path, first_string, second_string=None, third_string=None):
    def clean_str(s):
        return s.replace('[', '').replace(']', '').replace("'", "").replace('"', '').strip() if s else ""
    first_string, second_string, third_string = map(clean_str, (first_string, second_string, third_string))
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    matches = []
    for row in ws.iter_rows(values_only=True):
        row_values = [str(cell).strip() for cell in row if cell is not None]
        if any(first_string in val for val in row_values):
            if second_string and not any(second_string in val for val in row_values): continue
            if third_string and not any(third_string in val for val in row_values): continue
            matches.append(row_values)
    wb.close()
    return matches

# Searches for two strings (uppercase) in any row of Excel
def search_System_Serial_excel(file_path, str1, str2):
    str1, str2 = (s.replace("[", "").replace("]", "").strip().upper() if s else "" for s in (str1, str2))
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        row_values = [clean_cell_value_upper(cell) for cell in row if cell is not None]
        if (any(str1 in val for val in row_values) and any(str2 in val for val in row_values)):
            wb.close()
            return True
    wb.close()
    return False

# Highlights rows in Excel based on value in a given column (PASS/FAIL)
def highlight_rows_by_pass_fail_Transducer_System_Serial(file_path, sheet_number=1, column_index=4):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.worksheets[sheet_number - 1]
    green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    red_fill = PatternFill(start_color="FF7F7F", end_color="FF7F7F", fill_type="solid")
    for row in ws.iter_rows(min_row=2):
        cell_value = str(row[column_index - 1].value).strip().upper() if row[column_index - 1].value else ""
        if cell_value == "PASS":
            for cell in row:
                if cell.value not in (None, ""):
                    cell.fill = green_fill
        elif cell_value == "FAIL":
            for cell in row:
                if cell.value not in (None, ""):
                    cell.fill = red_fill
    wb.save(file_path)
    wb.close()

# Searches for three uppercase strings in a row in transducer sheet
def search_Transducer_excel(file_path, str1, str2, str3):
    str1, str2, str3 = (clean_cell_value_upper(s) for s in (str1, str2, str3))
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        row_values = [clean_cell_value_upper(cell) for cell in row if cell is not None]
        if (any(str1 in val for val in row_values) and
            any(str2 in val for val in row_values) and
            any(str3 in val for val in row_values)):
            wb.close()
            return True
    wb.close()
    return False

# Writes data_list to a row in Excel, ensures heading in row 1
def write_list_with_heading_Transducer(Excel_file_name, sheet_number, row_number, heading, data_list):
    if os.path.exists(Excel_file_name):
        wb = load_workbook(Excel_file_name)
    else:
        wb = Workbook()
    ws = get_or_create_sheet(wb, sheet_number)
    # Write heading if not already present
    existing_heading = [ws.cell(row=1, column=col).value for col in range(1, len(heading) + 1)]
    if existing_heading != list(heading):
        for col, value in enumerate(heading, start=1):
            ws.cell(row=1, column=col, value=value)
    for col, value in enumerate(data_list, start=1):
        ws.cell(row=row_number, column=col, value=value)
    wb.save(Excel_file_name)
    wb.close()


 
# Main orchestrator to write all outputs to Excel
def Excel_output_Entry(TestData_Results, TestProcedure_Results, SystemUsedRaw, FailureSummaryRaw_tuple, TestEquipmentRaw, Folder_Path):
    # Prepares and writes all summary, plan, system and transducer outputs
    Common_sheet = ['Filename', 'FileType']
    compare_Data_procedure_sheet_Dict = ['Initials', 'Start Date', 'Test Procedure Name', 'Baseline']
    Remaining_Sheet = ['Stop Date', 'Actual Hours']
    compare_Data_procedure_sheet_Not_Dict = ['Issue_Heading']
    check_procedure_sheet = ['Intials_Issues', 'FormField_Issues', 'Intials_Mismatches_Footer','Date_Mismatches_Footer', 'EndKeywordPresent']
    ExPlan_sheet = ['Test Procedure Name', 'Test Procedure View', 'Product', 'Subproduct','Test Folder Number', 'Sections Performed', 'MaketherowGreenP']
    ExSystemUsed_Sheet = ['System Serial Number', 'Cart version', 'MaketheRowGreenS']

    FailureSummaryRaw_list = [tuple(item) for item in FailureSummaryRaw_tuple]
    print("FAILURE SUMMARY LIST ")
    print(FailureSummaryRaw_list)
    SystemUsedRaw_list = [tuple(item) for item in SystemUsedRaw]

    Sheet1_Printed_Keys = Common_sheet + compare_Data_procedure_sheet_Dict + Remaining_Sheet
    Sheet2_Printed_Keys = compare_Data_procedure_sheet_Not_Dict + check_procedure_sheet
    Sheet3_Printed_Keys = ExPlan_sheet
    Sheet4_Printed_Keys = ExSystemUsed_Sheet
   

    save_results_to_excel(os.path.join(Folder_Path, "Quick_Summary.xlsx"), Sheet1_Printed_Keys, TestData_Results, TestProcedure_Results, start_row=2, sheet_number=1)
    compare_rows_sheet1_sheet2(os.path.join(Folder_Path, "Quick_Summary.xlsx"), sheet_number=1, row1=2, row2=3, start_col=3, last_col=6)

    save_results_to_excel(os.path.join(Folder_Path, "Quick_Summary.xlsx"), Sheet2_Printed_Keys, TestData_Results, TestProcedure_Results, start_row=2, sheet_number=2)
    write_list_to_cell(os.path.join(Folder_Path, "Quick_Summary.xlsx"), sheet_number=2, row=4, col=1, data_list=FailureSummaryRaw_list)
    compare_rows_sheet1_sheet2(os.path.join(Folder_Path, "Quick_Summary.xlsx"), sheet_number=2, row1=3, row2=4, start_col=1, last_col=1)

    TestData_Results_System = dict(TestData_Results)
    TestData_Results_System["System Serial Number"] = str(SystemUsedRaw_list[0][0]) if SystemUsedRaw_list else "N/A"
    TestData_Results_System["Sections Performed"] = [    item[1] if len(item) > 1 else None   # or "" / "N/A"
    for item in SystemUsedRaw_list
   ]
    MaketherowGreenP = search_Plan_excel("Checklist/PLAN - Sample.xlsx",TestData_Results_System.get("Test Procedure Name", ""),TestData_Results_System.get("Test Procedure View", ""),str(TestData_Results_System.get("Sections Performed", "")))
    TestData_Results_System["MaketherowGreenP"] = "PASS" if MaketherowGreenP else "FAIL"
    save_list_to_excel_single(os.path.join(Folder_Path, "Quick_Summary.xlsx"), Sheet3_Printed_Keys, TestData_Results_System, start_row=2, sheet_number=3)
    highlight_rows_by_pass_fail_Transducer_System_Serial(os.path.join(Folder_Path, "Quick_Summary.xlsx"), sheet_number=3, column_index=7)

    TestData_Results_System["Cart version"] = TestData_Results_System.get("Product", "") + " " + TestData_Results_System.get("Subproduct", "")
    Found = search_System_Serial_excel("Checklist/System Serial Numbers - Sample.xlsx",TestData_Results_System["System Serial Number"],TestData_Results_System["Cart version"])
    TestData_Results_System["MaketheRowGreenS"] = "PASS" if Found else "FAIL"
    save_list_to_excel_single(os.path.join(Folder_Path, "Quick_Summary.xlsx"), Sheet4_Printed_Keys, TestData_Results_System, start_row=2, sheet_number=4)
    highlight_rows_by_pass_fail_Transducer_System_Serial(os.path.join(Folder_Path, "Quick_Summary.xlsx"), sheet_number=4, column_index=3)

    for i, row in enumerate(TestEquipmentRaw, start=2):
        if len(row) == 3:
            var1, var2, var3 = row
            result = search_Transducer_excel("Checklist/Transducers - Sample.xlsx", var1, var2, TestData_Results.get("Product", ""))
            MaketheRowGreenT = "PASS" if result else "FAIL"
            headings = ["Probe Name", "Serial No", "Status/Date", "Product", "Pass/Fail"]
            write_list_with_heading_Transducer(os.path.join(Folder_Path, "Quick_Summary.xlsx"), sheet_number=5, row_number=i, heading=headings,data_list=[var1, var2, var3, TestData_Results.get("Product", ""), MaketheRowGreenT])
            highlight_rows_by_pass_fail_Transducer_System_Serial(os.path.join(Folder_Path, "Quick_Summary.xlsx"), sheet_number=5, column_index=5)


